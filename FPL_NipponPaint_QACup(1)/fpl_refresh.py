"""
FPL Auto-Refresh — Nippon Paint QA Cup
=======================================
Jalankan script ini untuk mengambil data terbaru dari FPL API
dan langsung memperbarui file Excel.

Cara pakai:
  - Windows : double-klik  run_refresh.bat
  - Mac/Linux: jalankan    python3 fpl_refresh.py
"""

import json
import os
import sys
import time
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path

# ── Konfigurasi ───────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
OUT_XLSX = BASE_DIR / "NipponPaint_QA_Cup_FPL.xlsx"

TEAMS = {
    "Sky Pool":      "6360719",
    "Giri-giri":     "5952186",
    "Sinar Muda FC": "6343832",
    "vnr":           "6301689",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json",
}

# ── Fungsi helper ─────────────────────────────────────────────────────────────
def log(msg, emoji=""):
    stamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{stamp}] {emoji}  {msg}" if emoji else f"[{stamp}]  {msg}")


def fetch_json(url, retries=3, delay=2):
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=15) as resp:
                return json.loads(resp.read().decode())
        except urllib.error.HTTPError as e:
            log(f"HTTP {e.code} saat akses {url} (attempt {attempt}/{retries})", "⚠️")
        except Exception as e:
            log(f"Error: {e} (attempt {attempt}/{retries})", "⚠️")
        if attempt < retries:
            time.sleep(delay)
    raise RuntimeError(f"Gagal mengambil data dari: {url}")


def get_current_gameweek():
    data = fetch_json("https://fantasy.premierleague.com/api/bootstrap-static/")
    for ev in data["events"]:
        if ev["is_current"]:
            return ev["id"]
    # fallback: ambil yang terakhir finished
    finished = [ev for ev in data["events"] if ev["finished"]]
    return finished[-1]["id"] if finished else 1


# ── Fetch picks & player data ────────────────────────────────────────────────
def fetch_picks_and_players(n_gw):
    """Fetch per-GW squad picks for all teams + player/club info."""
    import time as _time

    # Bootstrap: player info
    log("Mengambil data pemain dari FPL ...", "📡")
    boot = fetch_json("https://fantasy.premierleague.com/api/bootstrap-static/")
    team_map = {t["id"]: t["name"] for t in boot["teams"]}
    players = {}
    for p in boot["elements"]:
        players[p["id"]] = {"name": p["web_name"], "team": team_map[p["team"]]}

    # GW live points per player
    gw_player_pts = {}
    for gw in range(1, n_gw + 1):
        url  = f"https://fantasy.premierleague.com/api/event/{gw}/live/"
        data = fetch_json(url)
        gw_player_pts[gw] = {el["id"]: el["stats"]["total_points"] for el in data["elements"]}
        _time.sleep(0.2)
        if gw % 5 == 0:
            log(f"  Live points: GW{gw}/{n_gw}", "")

    # Picks per team per GW
    all_picks = {}
    for team_name, entry_id in TEAMS.items():
        all_picks[team_name] = {}
        for gw in range(1, n_gw + 1):
            url  = f"https://fantasy.premierleague.com/api/entry/{entry_id}/event/{gw}/picks/"
            try:
                data = fetch_json(url)
                all_picks[team_name][gw] = data["picks"]
            except Exception:
                all_picks[team_name][gw] = []
            _time.sleep(0.2)
        log(f"  Picks fetched: {team_name}", "")

    return players, gw_player_pts, all_picks


# ── Fetch data per tim ────────────────────────────────────────────────────────
def fetch_all_teams():
    results = {}
    for team_name, entry_id in TEAMS.items():
        log(f"Mengambil data: {team_name} ...", "📡")
        url  = f"https://fantasy.premierleague.com/api/entry/{entry_id}/history/"
        data = fetch_json(url)
        results[team_name] = data
        time.sleep(0.5)  # jeda kecil agar tidak kena rate-limit
    return results


# ── Build Excel ───────────────────────────────────────────────────────────────
def build_excel(all_data, current_gw):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule

    TEAM_LIST = list(TEAMS.keys())

    TEAM_HEX = {
        "Sky Pool":      "1565C0",
        "Giri-giri":     "C62828",
        "Sinar Muda FC": "2E7D32",
        "vnr":           "E65100",
    }
    TEAM_LIGHT = {
        "Sky Pool":      "DBEAFE",
        "Giri-giri":     "FEE2E2",
        "Sinar Muda FC": "DCFCE7",
        "vnr":           "FFF7ED",
    }
    CHIP_NAMES = {
        "wildcard": "Wildcard",
        "bboost":   "Bench Boost",
        "freehit":  "Free Hit",
        "3xc":      "Triple Captain",
    }
    CHIP_COLORS = {
        "Wildcard":       ("E3F2FD", "0D47A1"),
        "Bench Boost":    ("F3E5F5", "4A148C"),
        "Free Hit":       ("E8F5E9", "1B5E20"),
        "Triple Captain": ("FFF8E1", "E65100"),
    }
    DARK_BG = "1A1A2E"
    MID_BG  = "16213E"
    WHITE   = "FFFFFF"
    RANK_FILL = {"1": "D4EDDA", "2": "FFF3CD", "3": "FFE0B2", "4": "F8D7DA"}
    RANK_FONT = {"1": "155724", "2": "856404", "3": "BF360C", "4": "721C24"}

    def fill(h):
        return PatternFill("solid", fgColor=h)

    def hdr_font(sz=10, bold=True, color=WHITE):
        return Font(name="Calibri", size=sz, bold=bold, color=color)

    def cell_font(sz=9, bold=False, color="000000"):
        return Font(name="Calibri", size=sz, bold=bold, color=color)

    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def _side(style="thin", color="DDDDDD"):
        return Side(style=style, color=color)

    def thin_border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)

    def thick_bottom():
        return Border(left=_side(), right=_side(), top=_side(),
                      bottom=_side("medium", "AAAAAA"))

    # Build GW data
    n_gw = current_gw
    gws  = list(range(1, n_gw + 1))

    gw_pts     = {}
    cumulative = {}
    for team in TEAM_LIST:
        current = all_data[team]["current"][:n_gw]
        # Gunakan total_points dari API (sudah dikurangi hit transfer)
        pts = [gw["points"] for gw in current]
        cum = [gw["total_points"] for gw in current]
        gw_pts[team]     = pts
        cumulative[team] = cum

    weekly_rank   = {t: [] for t in TEAM_LIST}
    weekly_winner = []
    weekly_last   = []
    for gw_i in range(n_gw):
        snap   = {t: cumulative[t][gw_i] for t in TEAM_LIST}
        ranked = sorted(snap, key=snap.get, reverse=True)
        weekly_winner.append(ranked[0])
        weekly_last.append(ranked[-1])
        for pos, team in enumerate(ranked):
            weekly_rank[team].append(pos + 1)

    win_count  = {t: weekly_winner.count(t) for t in TEAM_LIST}
    last_count = {t: weekly_last.count(t)   for t in TEAM_LIST}

    wb = Workbook()

    # ── SHEET 1: Dashboard ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 38
    ws.merge_cells("B2:M2")
    ws["B2"] = f"NIPPON PAINT QA CUP  —  FPL Stats  |  Season 2025/26  |  s/d GW{n_gw}"
    ws["B2"].font = hdr_font(15); ws["B2"].fill = fill(DARK_BG); ws["B2"].alignment = center()

    ws.row_dimensions[3].height = 20
    ws.merge_cells("B3:M3")
    ws["B3"] = f"Terakhir diperbarui: {datetime.now().strftime('%d %b %Y %H:%M')} WIB  |  Source: fantasy.premierleague.com"
    ws["B3"].font = Font(name="Calibri", size=9, italic=True, color="AAAAAA")
    ws["B3"].fill = fill(DARK_BG); ws["B3"].alignment = center()
    ws.row_dimensions[4].height = 8

    # Nav
    from openpyxl.worksheet.hyperlink import Hyperlink
    nav = [("Dashboard","Dashboard"),("Klasemen GW","Klasemen per GW"),
           ("Poin GW","Poin per GW"),("Rank GW","Rank per GW"),
           ("Chip","Chip Usage"),("Raw Data","Raw Data")]
    ws.row_dimensions[5].height = 22
    for (lbl, sht), col in zip(nav, [2,4,6,8,10,12]):
        c = ws.cell(5, col, lbl)
        c.hyperlink = Hyperlink(ref=c.coordinate, location=f"'{sht}'!A1")
        c.font = Font(name="Calibri", size=9, bold=True, color=WHITE, underline="single")
        c.fill = fill("0F3460"); c.alignment = center()
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.row_dimensions[6].height = 10

    # Standings card
    ws.row_dimensions[7].height = 22
    for lbl, col in zip(["#","Tim","GW Pts","Total","GW #1","GW Juru Kunci"],
                         [2,3,5,7,9,11]):
        c = ws.cell(7, col, lbl)
        c.font = hdr_font(10); c.fill = fill("0A0A23")
        c.alignment = center(); c.border = thick_bottom()

    for i, team in enumerate(sorted(TEAM_LIST, key=lambda t: cumulative[t][-1], reverse=True)):
        row = 8 + i
        ws.row_dimensions[row].height = 22
        vals = [i+1, team, gw_pts[team][-1], cumulative[team][-1],
                win_count[team], last_count[team]]
        for val, col in zip(vals, [2,3,5,7,9,11]):
            c = ws.cell(row, col, val)
            c.fill = fill(TEAM_LIGHT[team]); c.border = thin_border(); c.alignment = center()
            is_total = (col == 7)
            c.font = Font(name="Calibri", size=10 if not is_total else 11,
                          bold=(col in [3,7]), color=TEAM_HEX[team])

    # Chart helper data
    cd = 20
    ws.cell(cd, 2, "Tim"); ws.cell(cd, 3, "GW #1"); ws.cell(cd, 4, "GW Juru Kunci")
    for i, t in enumerate(TEAM_LIST):
        ws.cell(cd+1+i, 2, t)
        ws.cell(cd+1+i, 3, win_count[t])
        ws.cell(cd+1+i, 4, last_count[t])

    bc = BarChart()
    bc.type = "bar"; bc.title = "GW sebagai #1 vs Juru Kunci"; bc.style = 10
    bc.width = 18; bc.height = 12
    bc.add_data(Reference(ws, min_col=3, max_col=4, min_row=cd, max_row=cd+len(TEAM_LIST)),
                titles_from_data=True)
    bc.set_categories(Reference(ws, min_col=2, min_row=cd+1, max_row=cd+len(TEAM_LIST)))
    ws.add_chart(bc, "B14")

    # ── SHEET 2: Klasemen per GW ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Klasemen per GW")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 3
    ws2.merge_cells("B2:L2")
    ws2["B2"] = "Klasemen per Gameweek — Pemimpin (#1) & Juru Kunci (#4)"
    ws2["B2"].font = hdr_font(13); ws2["B2"].fill = fill(DARK_BG); ws2["B2"].alignment = center()
    ws2.row_dimensions[2].height = 32; ws2.row_dimensions[3].height = 8

    hdrs = ["GW","Pemimpin (#1)","Total #1","Juru Kunci (#4)","Total #4",
            *[f"{t}" for t in TEAM_LIST]]
    col_ws2 = [6,18,12,18,12,13,13,13,13]
    ws2.row_dimensions[4].height = 22
    for j,(h,w) in enumerate(zip(hdrs, col_ws2)):
        c = ws2.cell(4, 2+j, h)
        c.font = hdr_font(9); c.fill = fill(DARK_BG)
        c.alignment = center(wrap=True); c.border = thick_bottom()
        ws2.column_dimensions[get_column_letter(2+j)].width = w

    for gw_i, gw in enumerate(gws):
        row = 5 + gw_i
        ws2.row_dimensions[row].height = 17
        winner = weekly_winner[gw_i]; loser = weekly_last[gw_i]
        row_data = [gw, winner, cumulative[winner][gw_i],
                    loser, cumulative[loser][gw_i],
                    *[cumulative[t][gw_i] for t in TEAM_LIST]]
        for j, val in enumerate(row_data):
            c = ws2.cell(row, 2+j, val)
            c.border = thin_border(); c.alignment = center()
            c.fill = fill("F8F9FA" if gw_i%2==0 else "FFFFFF")
            c.font = cell_font()
        # Winner highlight
        for col_off, clr_key in [(1, winner),(2, winner)]:
            c = ws2.cell(row, 2+col_off)
            c.fill = fill(TEAM_LIGHT[winner])
            c.font = Font(name="Calibri", size=9, bold=True, color=TEAM_HEX[winner])
        # Loser highlight
        for col_off in [3, 4]:
            c = ws2.cell(row, 2+col_off)
            c.fill = fill("FFF0F0")
            c.font = Font(name="Calibri", size=9, bold=True, color="CC0000")
        # Team totals
        for ti, team in enumerate(TEAM_LIST):
            c = ws2.cell(row, 7+ti)
            c.fill = fill(TEAM_LIGHT[team])
            c.font = cell_font(9, color=TEAM_HEX[team])
            c.number_format = "#,##0"

    ws2.freeze_panes = "B5"

    # Summary
    sr = 5 + n_gw + 2
    ws2.merge_cells(f"B{sr}:D{sr}")
    ws2[f"B{sr}"] = "RINGKASAN"
    ws2[f"B{sr}"].font = hdr_font(10); ws2[f"B{sr}"].fill = fill(DARK_BG); ws2[f"B{sr}"].alignment = center()
    sr += 1
    for j, lbl in enumerate(["Tim","GW Sebagai #1","GW Sebagai #4"]):
        c = ws2.cell(sr, 2+j, lbl)
        c.font = hdr_font(9); c.fill = fill(MID_BG); c.alignment = center(); c.border = thick_bottom()
    for i, team in enumerate(sorted(TEAM_LIST, key=lambda t: win_count[t], reverse=True)):
        r = sr+1+i
        ws2.row_dimensions[r].height = 18
        for j, (val, tfill, tfont) in enumerate([
            (team,          TEAM_LIGHT[team], TEAM_HEX[team]),
            (win_count[team], "D4EDDA",       "155724"),
            (last_count[team],"F8D7DA",       "CC0000"),
        ]):
            c = ws2.cell(r, 2+j, val)
            c.fill = fill(tfill); c.font = Font(name="Calibri",size=10,bold=True,color=tfont)
            c.alignment = center(); c.border = thin_border()

    # ── SHEET 3: Poin per GW ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Poin per GW")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 3
    ws3.merge_cells("B2:G2")
    ws3["B2"] = "Poin per Gameweek — Per Tim  (hijau=tertinggi, merah=terendah)"
    ws3["B2"].font = hdr_font(13); ws3["B2"].fill = fill(DARK_BG); ws3["B2"].alignment = center()
    ws3.row_dimensions[2].height = 32; ws3.row_dimensions[3].height = 8; ws3.row_dimensions[4].height = 22

    for j, h in enumerate(["GW", *TEAM_LIST]):
        c = ws3.cell(4, 2+j, h)
        c.font = hdr_font(9, color=WHITE if j==0 else TEAM_HEX[TEAM_LIST[j-1]])
        c.fill = fill(DARK_BG if j==0 else TEAM_LIGHT[TEAM_LIST[j-1]])
        c.alignment = center(); c.border = thick_bottom()
        ws3.column_dimensions[get_column_letter(2+j)].width = 16

    for gw_i, gw in enumerate(gws):
        row = 5 + gw_i
        ws3.row_dimensions[row].height = 17
        c = ws3.cell(row, 2, gw)
        c.font = cell_font(9, True); c.alignment = center()
        c.fill = fill("F0F0F8"); c.border = thin_border()
        max_p = max(gw_pts[t][gw_i] for t in TEAM_LIST)
        min_p = min(gw_pts[t][gw_i] for t in TEAM_LIST)
        for ti, team in enumerate(TEAM_LIST):
            p = gw_pts[team][gw_i]
            c = ws3.cell(row, 3+ti, p)
            c.border = thin_border(); c.alignment = center()
            if p == max_p:
                c.fill = fill("D4EDDA"); c.font = Font(name="Calibri",size=9,bold=True,color="155724")
            elif p == min_p:
                c.fill = fill("F8D7DA"); c.font = Font(name="Calibri",size=9,bold=True,color="721C24")
            else:
                c.fill = fill("FFFFFF" if gw_i%2==0 else "F8F9FA"); c.font = cell_font()

    ws3.freeze_panes = "B5"
    tot_row = 5 + n_gw
    ws3.row_dimensions[tot_row].height = 20
    c = ws3.cell(tot_row, 2, "TOTAL")
    c.font = hdr_font(9); c.fill = fill(DARK_BG); c.alignment = center(); c.border = thin_border()
    for ti, team in enumerate(TEAM_LIST):
        col_l = get_column_letter(3+ti)
        c = ws3.cell(tot_row, 3+ti)
        c.value = f"=SUM({col_l}5:{col_l}{tot_row-1})"
        c.number_format = "#,##0"
        c.font = Font(name="Calibri",size=10,bold=True,color=TEAM_HEX[team])
        c.fill = fill(TEAM_LIGHT[team]); c.alignment = center(); c.border = thin_border()
        ws3.conditional_formatting.add(f"{col_l}5:{col_l}{tot_row-1}",
            ColorScaleRule(start_type="min",start_color="FFCCCC",
                           mid_type="percentile",mid_value=50,mid_color="FFFFFF",
                           end_type="max",end_color="CCFFCC"))

    # Cumulative helper + line chart
    csr = tot_row + 3
    ws3.cell(csr, 2, "GW")
    for ti, t in enumerate(TEAM_LIST): ws3.cell(csr, 3+ti, t)
    for gw_i, gw in enumerate(gws):
        r = csr+1+gw_i
        ws3.cell(r, 2, gw)
        for ti, t in enumerate(TEAM_LIST): ws3.cell(r, 3+ti, cumulative[t][gw_i])
    cer = csr + n_gw
    lc = LineChart()
    lc.title = "Akumulasi Poin per Gameweek"; lc.style = 10
    lc.y_axis.title = "Total Poin"; lc.x_axis.title = "Gameweek"
    lc.width = 24; lc.height = 14
    lc.add_data(Reference(ws3, min_col=3, max_col=3+len(TEAM_LIST)-1,
                          min_row=csr, max_row=cer), titles_from_data=True)
    lc.set_categories(Reference(ws3, min_col=2, min_row=csr+1, max_row=cer))
    ws3.add_chart(lc, f"B{cer+3}")

    # ── SHEET 4: Rank per GW ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Rank per GW")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 3
    ws4.merge_cells("B2:G2")
    ws4["B2"] = "Posisi / Rank per Gameweek  (1=Pemimpin, 4=Juru Kunci)"
    ws4["B2"].font = hdr_font(13); ws4["B2"].fill = fill(DARK_BG); ws4["B2"].alignment = center()
    ws4.row_dimensions[2].height = 32; ws4.row_dimensions[3].height = 8; ws4.row_dimensions[4].height = 22

    for j, h in enumerate(["GW", *TEAM_LIST]):
        c = ws4.cell(4, 2+j, h)
        c.font = hdr_font(9, color=WHITE if j==0 else TEAM_HEX[TEAM_LIST[j-1]])
        c.fill = fill(DARK_BG if j==0 else TEAM_LIGHT[TEAM_LIST[j-1]])
        c.alignment = center(); c.border = thick_bottom()
        ws4.column_dimensions[get_column_letter(2+j)].width = 16

    for gw_i, gw in enumerate(gws):
        row = 5+gw_i; ws4.row_dimensions[row].height = 17
        c = ws4.cell(row, 2, gw)
        c.font = cell_font(9,True); c.alignment = center()
        c.fill = fill("F0F0F8"); c.border = thin_border()
        for ti, team in enumerate(TEAM_LIST):
            rank = weekly_rank[team][gw_i]
            c = ws4.cell(row, 3+ti, rank)
            c.border = thin_border(); c.alignment = center()
            c.fill = fill(RANK_FILL[str(rank)])
            c.font = Font(name="Calibri",size=9,bold=True,color=RANK_FONT[str(rank)])
    ws4.freeze_panes = "B5"

    sr4 = 5+n_gw+2
    ws4.merge_cells(f"B{sr4}:D{sr4}")
    ws4[f"B{sr4}"] = "Rekap: Berapa Kali Jadi #1 dan #4"
    ws4[f"B{sr4}"].font = hdr_font(10); ws4[f"B{sr4}"].fill = fill(DARK_BG); ws4[f"B{sr4}"].alignment = center()
    sr4 += 1
    for j, lbl in enumerate(["Tim","Kali Jadi #1","Kali Jadi #4"]):
        c = ws4.cell(sr4, 2+j, lbl)
        c.font = hdr_font(9); c.fill = fill(MID_BG); c.alignment = center(); c.border = thick_bottom()
    for i, team in enumerate(sorted(TEAM_LIST, key=lambda t: win_count[t], reverse=True)):
        r = sr4+1+i; ws4.row_dimensions[r].height = 18
        for j, (val, tfill, tfont) in enumerate([
            (team, TEAM_LIGHT[team], TEAM_HEX[team]),
            (win_count[team], "D4EDDA", "155724"),
            (last_count[team],"F8D7DA", "CC0000"),
        ]):
            c = ws4.cell(r, 2+j, val)
            c.fill = fill(tfill); c.font = Font(name="Calibri",size=10,bold=True,color=tfont)
            c.alignment = center(); c.border = thin_border()

    # ── SHEET 5: Chip Usage ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Chip Usage")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 3
    ws5.merge_cells("B2:F2")
    ws5["B2"] = "Chip Usage — Kapan & Apa Chip yang Dipakai"
    ws5["B2"].font = hdr_font(13); ws5["B2"].fill = fill(DARK_BG); ws5["B2"].alignment = center()
    ws5.row_dimensions[2].height = 32; ws5.row_dimensions[3].height = 8; ws5.row_dimensions[4].height = 22
    for j, (h, w) in enumerate(zip(["Tim","Chip","GW Dipakai","Poin GW Itu","Total Saat Itu"],
                                    [20, 18, 12, 14, 18])):
        c = ws5.cell(4, 2+j, h)
        c.font = hdr_font(9); c.fill = fill(DARK_BG); c.alignment = center(); c.border = thick_bottom()
        ws5.column_dimensions[get_column_letter(2+j)].width = w

    chip_row = 5
    for team in TEAM_LIST:
        chips = sorted(all_data[team].get("chips",[]), key=lambda x: x["event"])
        for ch in chips:
            gw_used = ch["event"]
            if gw_used > n_gw:
                continue
            chip_name = CHIP_NAMES.get(ch["name"], ch["name"])
            bg, fg = CHIP_COLORS.get(chip_name, ("FFFFFF","000000"))
            pts_gw   = gw_pts[team][gw_used-1]
            total_gw = cumulative[team][gw_used-1]
            for j, val in enumerate([team, chip_name, gw_used, pts_gw, total_gw]):
                c = ws5.cell(chip_row, 2+j, val)
                c.fill = fill(TEAM_LIGHT[team] if j==0 else bg)
                c.font = Font(name="Calibri",size=9,bold=(j<2),
                              color=TEAM_HEX[team] if j==0 else fg)
                c.alignment = center(); c.border = thin_border()
            chip_row += 1

    # ── SHEET 6: Raw Data ─────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Raw Data")
    ws6.row_dimensions[1].height = 22
    for j, (h, w) in enumerate(zip(
        ["GW","Tim","Poin GW","Total Kumulatif","Rank di Liga","Poin di Bench","Transfer"],
        [6, 18, 12, 18, 14, 16, 12]
    )):
        c = ws6.cell(1, 1+j, h)
        c.font = hdr_font(9); c.fill = fill(DARK_BG); c.alignment = center(); c.border = thick_bottom()
        ws6.column_dimensions[get_column_letter(1+j)].width = w

    raw_row = 2
    for team in TEAM_LIST:
        for gw_data in all_data[team]["current"][:n_gw]:
            gw_n = gw_data["event"]
            row_vals = [
                gw_n, team, gw_data["points"], gw_data["total_points"],
                weekly_rank[team][gw_n-1],
                gw_data.get("points_on_bench",0),
                gw_data.get("event_transfers",0),
            ]
            for j, val in enumerate(row_vals):
                c = ws6.cell(raw_row, 1+j, val)
                c.font = cell_font(); c.alignment = center(); c.border = thin_border()
                if j == 1:
                    c.fill = fill(TEAM_LIGHT[team])
                    c.font = Font(name="Calibri",size=9,color=TEAM_HEX[team])
                elif raw_row % 2 == 0:
                    c.fill = fill("F8F9FA")
            raw_row += 1

    ws6.auto_filter.ref = f"A1:{get_column_letter(7)}{raw_row-1}"
    ws6.freeze_panes = "A2"
    ws6.cell(raw_row+1, 1,
             f"Source: fantasy.premierleague.com  |  League: 1429930  |  "
             f"Refreshed: {datetime.now().strftime('%d %b %Y %H:%M')}")

    wb.save(str(OUT_XLSX))
    log(f"File Excel disimpan: {OUT_XLSX}", "✅")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  FPL Auto-Refresh — Nippon Paint QA Cup")
    print("=" * 55)
    log("Memeriksa gameweek aktif ...", "🔍")
    try:
        current_gw = get_current_gameweek()
        log(f"Gameweek aktif: GW{current_gw}", "📅")
    except Exception as e:
        log(f"Tidak bisa deteksi GW, pakai semua data tersedia. ({e})", "⚠️")
        current_gw = 38  # fallback max

    log("Mengambil data semua tim dari FPL API ...", "⬇️")
    try:
        all_data = fetch_all_teams()
    except RuntimeError as e:
        log(str(e), "❌")
        input("\nTekan Enter untuk keluar ...")
        sys.exit(1)

    # Clamp ke GW yang datanya tersedia
    actual_gw = min(current_gw, len(all_data["Sky Pool"]["current"]))
    log(f"Data tersedia hingga GW{actual_gw}", "📊")

    log("Mengambil data picks & pemain (untuk Distribusi Klub) ...", "⬇️")
    try:
        players, gw_player_pts, all_picks = fetch_picks_and_players(actual_gw)
    except Exception as e:
        log(f"Gagal fetch picks: {e} — sheet Distribusi Klub dilewati.", "⚠️")
        players, gw_player_pts, all_picks = None, None, None

    log("Membuat file Excel ...", "📝")
    build_excel(all_data, actual_gw)

    if players is not None:
        log("Menambahkan sheet Distribusi Klub ...", "📝")
        build_club_sheets(players, gw_player_pts, all_picks, actual_gw)

    print()
    print("=" * 55)
    print(f"  SELESAI! File diperbarui s/d GW{actual_gw}")
    print(f"  {OUT_XLSX}")
    print("=" * 55)
    input("\nTekan Enter untuk menutup jendela ini ...")


# ── Build Club Distribution Sheets ──────────────────────────────────────────
def build_club_sheets(players, gw_player_pts, all_picks, n_gw):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from collections import defaultdict

    TEAM_LIST  = list(TEAMS.keys())
    TEAM_HEX   = {"Sky Pool":"1565C0","Giri-giri":"C62828","Sinar Muda FC":"2E7D32","vnr":"E65100"}
    TEAM_LIGHT = {"Sky Pool":"DBEAFE","Giri-giri":"FEE2E2","Sinar Muda FC":"DCFCE7","vnr":"FFF7ED"}
    DARK_BG = "1A1A2E"; MID_BG = "16213E"; WHITE = "FFFFFF"
    ALL_CLUBS = sorted(set(p["team"] for p in players.values()))

    def fill(h):  return PatternFill("solid", fgColor=h)
    def center(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def left(): return Alignment(horizontal="left", vertical="center", indent=1)
    def _side(s="thin", c="DDDDDD"): return Side(style=s, color=c)
    def thin_border(): s=_side(); return Border(left=s,right=s,top=s,bottom=s)
    def thick_bottom():
        return Border(left=_side(),right=_side(),top=_side(),bottom=_side("medium","AAAAAA"))
    def hdr_font(sz=10, bold=True, color=WHITE):
        return Font(name="Calibri", size=sz, bold=bold, color=color)
    def cell_font(sz=9, bold=False, color="111111"):
        return Font(name="Calibri", size=sz, bold=bold, color=color)

    # Build per-GW club data
    gw_club_data  = {}
    gw_top_scorer = {}
    for gw in range(1, n_gw + 1):
        gw_club_data[gw]  = {}
        gw_top_scorer[gw] = {}
        for team in TEAM_LIST:
            picks = all_picks[team].get(gw, [])
            club_players = defaultdict(list)
            top_name, top_pts = "", 0
            for pick in picks:
                pid = pick["element"]
                if pid not in players: continue
                pname = players[pid]["name"]
                pteam = players[pid]["team"]
                club_players[pteam].append(pname)
                pts = gw_player_pts.get(gw, {}).get(pid, 0)
                if pts > top_pts:
                    top_pts = pts; top_name = pname
            gw_club_data[gw][team]  = dict(club_players)
            gw_top_scorer[gw][team] = (top_name, top_pts)

    # Accumulation
    accum_club = {team: defaultdict(int) for team in TEAM_LIST}
    for gw in range(1, n_gw + 1):
        for team in TEAM_LIST:
            for club, names in gw_club_data[gw][team].items():
                accum_club[team][club] += len(names)
    total_slots = {team: sum(accum_club[team].values()) for team in TEAM_LIST}

    wb = load_workbook(str(OUT_XLSX))
    for sname in ["Distribusi Klub GW", "Distribusi Klub Akumulasi"]:
        if sname in wb.sheetnames: del wb[sname]

    # ── Sheet A: per GW ──────────────────────────────────────────────────────
    ws_a = wb.create_sheet("Distribusi Klub GW")
    ws_a.sheet_view.showGridLines = False
    ws_a.column_dimensions["A"].width = 3
    ws_a.row_dimensions[1].height = 8
    ws_a.row_dimensions[2].height = 34
    ws_a.merge_cells("B2:R2")
    ws_a["B2"] = "Distribusi Klub per Gameweek — Filter kolom GW untuk pilih gameweek"
    ws_a["B2"].font = hdr_font(13); ws_a["B2"].fill = fill(DARK_BG); ws_a["B2"].alignment = center()
    ws_a.row_dimensions[3].height = 8

    hdr_row = 4
    ws_a.row_dimensions[hdr_row].height = 30
    hdrs = ["GW","Klub",
            "Sky Pool\n(Jml)","Sky Pool\n(Pemain)",
            "Giri-giri\n(Jml)","Giri-giri\n(Pemain)",
            "Sinar Muda\n(Jml)","Sinar Muda\n(Pemain)",
            "vnr\n(Jml)","vnr\n(Pemain)",
            "Top Scorer\nSky Pool","Top Scorer\nGiri-giri",
            "Top Scorer\nSinar Muda","Top Scorer\nvnr"]
    hdr_fills=[DARK_BG,DARK_BG,"1565C0","1565C0","C62828","C62828",
               "2E7D32","2E7D32","E65100","E65100","1565C0","C62828","2E7D32","E65100"]
    col_widths_a={2:6,3:16,4:10,5:28,6:10,7:28,8:14,9:28,10:8,11:28,12:20,13:20,14:20,15:20}
    for col,w in col_widths_a.items():
        ws_a.column_dimensions[get_column_letter(col)].width = w
    for j,(h,hf) in enumerate(zip(hdrs,hdr_fills)):
        c = ws_a.cell(hdr_row, 2+j, h)
        c.font = hdr_font(8); c.fill = fill(hf)
        c.alignment = center(wrap=True); c.border = thick_bottom()

    data_row = hdr_row + 1
    for gw in range(1, n_gw + 1):
        for club in ALL_CLUBS:
            ws_a.row_dimensions[data_row].height = 16
            row_fill = fill("F8F9FA") if gw%2==0 else fill("FFFFFF")
            c = ws_a.cell(data_row, 2, gw)
            c.font=cell_font(9,True); c.fill=fill("F0F0F8"); c.alignment=center(); c.border=thin_border()
            c = ws_a.cell(data_row, 3, club)
            c.font=cell_font(9); c.fill=row_fill; c.alignment=left(); c.border=thin_border()
            for ti, team in enumerate(TEAM_LIST):
                names = gw_club_data[gw][team].get(club, [])
                cnt   = len(names)
                c_cnt = ws_a.cell(data_row, 4+ti*2, cnt if cnt>0 else "")
                c_cnt.font=Font(name="Calibri",size=9,bold=(cnt>0),color=TEAM_HEX[team] if cnt>0 else "CCCCCC")
                c_cnt.fill=fill(TEAM_LIGHT[team]) if cnt>0 else row_fill
                c_cnt.alignment=center(); c_cnt.border=thin_border()
                c_names=ws_a.cell(data_row,5+ti*2,", ".join(names) if names else "-")
                c_names.font=Font(name="Calibri",size=8,color=TEAM_HEX[team] if names else "CCCCCC")
                c_names.fill=fill(TEAM_LIGHT[team]) if names else row_fill
                c_names.alignment=Alignment(horizontal="left",vertical="center",indent=1)
                c_names.border=thin_border()
            if club == ALL_CLUBS[0]:
                for ti, team in enumerate(TEAM_LIST):
                    top_name, top_pts = gw_top_scorer[gw][team]
                    c=ws_a.cell(data_row,12+ti,f"{top_name} ({top_pts}pts)" if top_name else "-")
                    c.font=Font(name="Calibri",size=8,bold=True,color=TEAM_HEX[team])
                    c.fill=fill(TEAM_LIGHT[team]); c.alignment=center(); c.border=thin_border()
            else:
                for ti in range(4):
                    c=ws_a.cell(data_row,12+ti,""); c.fill=row_fill; c.border=thin_border()
            data_row += 1

    ws_a.auto_filter.ref = f"B{hdr_row}:{get_column_letter(15)}{data_row-1}"
    ws_a.freeze_panes = f"B{hdr_row+1}"
    ws_a.cell(data_row+1,2,"Tip: Gunakan filter di kolom GW untuk melihat data per gameweek").font=\
        Font(name="Calibri",size=9,italic=True,color="888888")

    # ── Sheet B: akumulasi ────────────────────────────────────────────────────
    ws_b = wb.create_sheet("Distribusi Klub Akumulasi")
    ws_b.sheet_view.showGridLines = False
    ws_b.column_dimensions["A"].width = 3
    ws_b.row_dimensions[1].height = 8; ws_b.row_dimensions[2].height = 34
    ws_b.merge_cells("B2:L2")
    ws_b["B2"] = f"Distribusi Klub Akumulasi — GW1 s/d GW{n_gw} (% dari total slot pemain)"
    ws_b["B2"].font=hdr_font(13); ws_b["B2"].fill=fill(DARK_BG); ws_b["B2"].alignment=center()
    ws_b.row_dimensions[3].height = 18
    ws_b.merge_cells("B3:L3")
    ws_b["B3"] = f"Total slot per tim: {n_gw} GW x 15 pemain = {n_gw*15} slot."
    ws_b["B3"].font=Font(name="Calibri",size=9,italic=True,color="AAAAAA")
    ws_b["B3"].fill=fill(DARK_BG); ws_b["B3"].alignment=center()
    ws_b.row_dimensions[4].height = 8

    hdr_b = 5
    ws_b.row_dimensions[hdr_b].height = 28
    b_cols=["Klub","Sky Pool\n(%)","Sky Pool\n(Slot)","Giri-giri\n(%)","Giri-giri\n(Slot)",
            "Sinar Muda\n(%)","Sinar Muda\n(Slot)","vnr\n(%)","vnr\n(Slot)"]
    b_fills=[DARK_BG,"1565C0","1565C0","C62828","C62828","2E7D32","2E7D32","E65100","E65100"]
    b_widths=[20,12,10,12,10,14,10,10,10]
    for j,(h,hf,w) in enumerate(zip(b_cols,b_fills,b_widths)):
        c=ws_b.cell(hdr_b,2+j,h)
        c.font=hdr_font(9); c.fill=fill(hf)
        c.alignment=center(wrap=True); c.border=thick_bottom()
        ws_b.column_dimensions[get_column_letter(2+j)].width=w

    club_totals={club:sum(accum_club[t][club] for t in TEAM_LIST) for club in ALL_CLUBS}
    sorted_clubs=sorted(ALL_CLUBS,key=lambda c:club_totals[c],reverse=True)
    data_b = hdr_b + 1
    for i,club in enumerate(sorted_clubs):
        row=data_b+i; ws_b.row_dimensions[row].height=18
        base=fill("F8F9FA") if i%2==0 else fill("FFFFFF")
        c=ws_b.cell(row,2,club)
        c.font=cell_font(10,True); c.fill=base; c.alignment=left(); c.border=thin_border()
        for ti,team in enumerate(TEAM_LIST):
            slots=accum_club[team][club]
            pct=slots/total_slots[team] if total_slots[team]>0 else 0
            c_pct=ws_b.cell(row,3+ti*2,pct)
            c_pct.number_format="0.0%"
            c_pct.font=Font(name="Calibri",size=10,bold=(slots>0),color=TEAM_HEX[team] if slots>0 else "CCCCCC")
            c_pct.fill=fill(TEAM_LIGHT[team]) if slots>0 else base
            c_pct.alignment=center(); c_pct.border=thin_border()
            c_slot=ws_b.cell(row,4+ti*2,slots if slots>0 else "")
            c_slot.font=Font(name="Calibri",size=9,color="888888")
            c_slot.fill=fill(TEAM_LIGHT[team]) if slots>0 else base
            c_slot.alignment=center(); c_slot.border=thin_border()

    tot_row=data_b+len(sorted_clubs); ws_b.row_dimensions[tot_row].height=22
    c=ws_b.cell(tot_row,2,"TOTAL")
    c.font=hdr_font(10); c.fill=fill(DARK_BG); c.alignment=center(); c.border=thin_border()
    for ti,team in enumerate(TEAM_LIST):
        c_pct=ws_b.cell(tot_row,3+ti*2,1.0)
        c_pct.number_format="0%"
        c_pct.font=Font(name="Calibri",size=10,bold=True,color=TEAM_HEX[team])
        c_pct.fill=fill(TEAM_LIGHT[team]); c_pct.alignment=center(); c_pct.border=thin_border()
        c_slot=ws_b.cell(tot_row,4+ti*2,total_slots[team])
        c_slot.font=Font(name="Calibri",size=9,bold=True,color="444444")
        c_slot.fill=fill(TEAM_LIGHT[team]); c_slot.alignment=center(); c_slot.border=thin_border()

    sum_row=tot_row+2; ws_b.row_dimensions[sum_row].height=22
    ws_b.merge_cells(f"B{sum_row}:J{sum_row}")
    ws_b[f"B{sum_row}"]="TOP 3 KLUB FAVORIT PER TIM"
    ws_b[f"B{sum_row}"].font=hdr_font(10); ws_b[f"B{sum_row}"].fill=fill(DARK_BG); ws_b[f"B{sum_row}"].alignment=center()
    sum_row+=1; ws_b.row_dimensions[sum_row].height=22
    for j,h in enumerate(["Tim","#1 Klub","%","#2 Klub","%","#3 Klub","%"]):
        c=ws_b.cell(sum_row,2+j,h)
        c.font=hdr_font(9); c.fill=fill(MID_BG); c.alignment=center(); c.border=thick_bottom()
    for ti,team in enumerate(TEAM_LIST):
        r=sum_row+1+ti; ws_b.row_dimensions[r].height=20
        top3=sorted(accum_club[team].items(),key=lambda x:x[1],reverse=True)[:3]
        c=ws_b.cell(r,2,team)
        c.font=Font(name="Calibri",size=10,bold=True,color=TEAM_HEX[team])
        c.fill=fill(TEAM_LIGHT[team]); c.alignment=center(); c.border=thin_border()
        for k,(club,slots) in enumerate(top3):
            pct=slots/total_slots[team]
            c_club=ws_b.cell(r,3+k*2,club)
            c_club.font=Font(name="Calibri",size=10,bold=True,color=TEAM_HEX[team])
            c_club.fill=fill(TEAM_LIGHT[team]); c_club.alignment=center(); c_club.border=thin_border()
            c_pct2=ws_b.cell(r,4+k*2,pct)
            c_pct2.number_format="0.0%"
            c_pct2.font=Font(name="Calibri",size=10,bold=True,color=TEAM_HEX[team])
            c_pct2.fill=fill(TEAM_LIGHT[team]); c_pct2.alignment=center(); c_pct2.border=thin_border()
    ws_b.freeze_panes=f"B{hdr_b+1}"

    wb.save(str(OUT_XLSX))
    log("Sheet Distribusi Klub selesai.", "✅")


if __name__ == "__main__":
    main()
